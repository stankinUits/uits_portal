import os

from django.conf import settings
from django.http import HttpResponse
from django.views import View

import openpyxl
from openpyxl.cell.cell import MergedCell

from parcing_data_from_excel.models import Student, OutputForParcingModuleGrade
import datetime


def _set_cell_value(ws, row: int, col: int, value):
    """
    Write `value` into ws.cell(row, col), even if that cell
    lives in a merged range (redirects to the top-left cell).
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                cell = ws.cell(row=merged.min_row, column=merged.min_col)
                break
    cell.value = value


class FillTemplateView(View):
    """
    GET  /api/export/fill-template/<group_name>/
    Fills the template for <group_name>, then writes it out into
    MEDIA_ROOT/output_module_template_for_groups/.
    If a file already exists, appends an incremental suffix:
      <group_name>.xlsx
      <group_name>_1.xlsx
      <group_name>_2.xlsx
      ...
    """
    TEMPLATE_SUBPATH = 'excel_files/template_for_module.xlsx'
    OUTPUT_DIR       = 'output_module_template_for_groups'
    START_ROW        = 11

    # Excel columns: 1=A, 2=B, 3=C, 4=D
    COL_FIRST   = 2  # B
    COL_MIDDLE  = 3  # C
    COL_LAST    = 4  # D

    def get(self, request, group_name):
        # 1) Load template
        tpl_path = os.path.join(settings.MEDIA_ROOT, self.TEMPLATE_SUBPATH)
        if not os.path.exists(tpl_path):
            return HttpResponse(f"Template not found: {tpl_path}", status=404)

        wb = openpyxl.load_workbook(tpl_path)

        # 2) Select the worksheet matching the group (or fallback to active)
        if group_name in wb.sheetnames:
            ws = wb[group_name]
        else:
            ws = wb.active

        # 3) Fetch students and fill starting at row 11
        students = (
            Student.objects
            .filter(group_name=group_name)
            .order_by('student_num_in_list')
        )

        # Fill discipline in C4:N4
        for col in range(3, 15):  # C=3, N=14
            _set_cell_value(ws, 4, col, group_name)  # C4:N4 discipline
        # Fill group_name in C5:N5 and teacher in C7:N7
        for col in range(3, 15):  # C=3, N=14
            _set_cell_value(ws, 5, col, group_name)      # C5:N5 group_name
            _set_cell_value(ws, 7, col, "Teacher")    # C7:N7 teacher
        for idx, stu in enumerate(students):
            row = self.START_ROW + idx
            _set_cell_value(ws, row, 1, idx + 1)  # A: student number
            _set_cell_value(ws, row, self.COL_FIRST,  stu.first_name)    # B: Имя
            _set_cell_value(ws, row, self.COL_MIDDLE, stu.middle_name)   # C: Отчество
            _set_cell_value(ws, row, self.COL_LAST,   stu.last_name)     # D: Фамилия

        # 4) Ensure output directory exists
        out_dir = os.path.join(settings.MEDIA_ROOT, self.OUTPUT_DIR)
        os.makedirs(out_dir, exist_ok=True)

        # 5) Choose a unique filename
        base = group_name
        ext = ".xlsx"
        filename = f"{base}{ext}"
        out_path = os.path.join(out_dir, filename)

        counter = 1
        while os.path.exists(out_path):
            filename = f"{base}_{counter}{ext}"
            out_path = os.path.join(out_dir, filename)
            counter += 1

        # 6) Save and respond
        wb.save(out_path)
        return HttpResponse(f"Saved: {out_path}", content_type="text/plain")


class ExportModuleGradeByTeacherDisciplineView(View):
    """
    GET /api/export/module-grade-by-teacher-discipline/
    For each unique (id_teachers_in_discipline, discipline) in OutputForParcingModuleGrade,
    copies the template, creates/fills a sheet for each group, and fills it with students.
    The file is named: module_grade_{id_teachers_in_discipline}_{discipline}_{date}.xlsx
    """
    TEMPLATE_SUBPATH = 'excel_files/template_for_module.xlsx'
    OUTPUT_DIR = 'output_module_grade_by_teacher_discipline'
    START_ROW = 11
    COL_FIRST = 2  # B
    COL_MIDDLE = 3 # C
    COL_LAST = 4   # D

    def get(self, request):
        import shutil
        # 1. Get all unique (id_teachers_in_discipline, discipline) pairs
        pairs = (
            OutputForParcingModuleGrade.objects
            .values_list('id_teachers_in_discipline', 'discipline')
            .distinct()
        )
        if not pairs:
            return HttpResponse("No id_teachers_in_discipline/discipline pairs found.", status=404)

        tpl_path = os.path.join(settings.MEDIA_ROOT, self.TEMPLATE_SUBPATH)
        if not os.path.exists(tpl_path):
            return HttpResponse(f"Template not found: {tpl_path}", status=404)

        out_dir = os.path.join(settings.MEDIA_ROOT, self.OUTPUT_DIR)
        os.makedirs(out_dir, exist_ok=True)
        
        # Delete all existing Excel files in the output directory
        existing_files = [f for f in os.listdir(out_dir) if f.endswith('.xlsx')]
        for file in existing_files:
            file_path = os.path.join(out_dir, file)
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Error deleting file {file}: {e}")
        
        # Also delete any existing zip files
        zip_files = [f for f in os.listdir(out_dir) if f.endswith('.zip')]
        for zip_file in zip_files:
            zip_path = os.path.join(out_dir, zip_file)
            try:
                os.remove(zip_path)
            except Exception as e:
                print(f"Error deleting zip file {zip_file}: {e}")
        
        today = datetime.date.today().strftime('%Y%m%d')
        files_created = []

        for id_teacher, discipline in pairs:
            # 2. Get all teachers for this id_teachers_in_discipline+discipline
            teachers = (
                OutputForParcingModuleGrade.objects
                .filter(id_teachers_in_discipline=id_teacher, discipline=discipline)
                .values_list('teacher', flat=True)
                .distinct()
            )
            teachers_str = ', '.join(teachers)
            # 3. Get all groups for this id_teachers_in_discipline+discipline
            groups = (
                OutputForParcingModuleGrade.objects
                .filter(id_teachers_in_discipline=id_teacher, discipline=discipline)
                .values_list('group_name', flat=True)
                .distinct()
            )
            if not groups:
                continue
            # 4. Load template for this file
            wb = openpyxl.load_workbook(tpl_path)
            
            # Get the first sheet name (which might be "Поток" or any other name)
            template_sheet_name = wb.sheetnames[0] if wb.sheetnames else None
            
            # Remove all sheets except the first one (template)
            while len(wb.sheetnames) > 1:
                wb.remove(wb[wb.sheetnames[1]])
            
            # Create sheets for each group
            sheets_created = []
            for group in groups:
                # Copy the template sheet and rename it to the group name
                ws = wb.copy_worksheet(wb.active)
                ws.title = group
                sheets_created.append(group)
                
                # Clear existing data in the copied sheet
                for row in ws.iter_rows(min_row=self.START_ROW, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None
                
                # Fill discipline in C4:N4
                for col in range(3, 15):
                    _set_cell_value(ws, 4, col, discipline)
                # Fill group_name in C5:N5 and teachers in C7:N7
                for col in range(3, 15):
                    _set_cell_value(ws, 5, col, group)
                    _set_cell_value(ws, 7, col, teachers_str)
                students = Student.objects.filter(group_name=group).order_by('student_num_in_list')
                for idx, stu in enumerate(students):
                    row = self.START_ROW + idx
                    _set_cell_value(ws, row, 1, idx + 1)
                    _set_cell_value(ws, row, self.COL_FIRST,  stu.first_name)
                    _set_cell_value(ws, row, self.COL_MIDDLE, stu.middle_name)
                    _set_cell_value(ws, row, self.COL_LAST,   stu.last_name)
            
            # Remove the original template sheet (e.g., "Поток")
            if template_sheet_name and template_sheet_name in wb.sheetnames and template_sheet_name not in groups:
                wb.remove(wb[template_sheet_name])
            
            # 5. Save file with new naming
            safe_id_teacher = str(id_teacher).replace(' ', '_').replace('/', '_').replace('\\', '_') if id_teacher else 'noid'
            safe_discipline = str(discipline).replace(' ', '_').replace('/', '_').replace('\\', '_')
            filename = f"module_grade_{safe_id_teacher}_{safe_discipline}_{today}.xlsx"
            out_path = os.path.join(out_dir, filename)
            
            wb.save(out_path)
            files_created.append(filename)
        if not files_created:
            return HttpResponse("No files created (no groups or students found).", status=404)
        return HttpResponse(f"Files created: {', '.join(files_created)}", content_type="text/plain")
